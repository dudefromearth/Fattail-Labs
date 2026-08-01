#!/usr/bin/env python3
"""Import Documents/0DTE-Trade-Log.xlsx (Log sheet) into member_trade_log_*.

Each spreadsheet row is one closed trade with Entry/Exit/Net P/L.
We materialize:
  - TO_OPEN fill on Open Date (DEBIT Entry)
  - TO_CLOSE fill on Close Date (CREDIT Exit, pnl_amount = Net P/L)
  - Synthetic butterfly legs from Type/Width/QTY/Asset (position detail was light)

  cd server && .venv/bin/python import_0dte_xlsx.py \\
      --xlsx ~/Documents/0DTE-Trade-Log.xlsx \\
      --email ernie@fattail.ai \\
      --also-clone-to dev-admin@labs.local,ernie@dudefromearth.com,erniev@mac.com
"""

from __future__ import annotations

import argparse
import re
from datetime import datetime, time
from decimal import Decimal
from pathlib import Path

import openpyxl

import db

ADAPTER = "xlsx_0dte_trade_log"
DEFAULT_XLSX = Path.home() / "Documents" / "0DTE-Trade-Log.xlsx"


def _dt(v) -> datetime | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return datetime.combine(v.date(), time(12, 0, 0))
    if isinstance(v, str):
        s = v.strip().replace("Feburary", "February")
        # "Oct 14,2025" → "Oct 14, 2025"
        s = re.sub(r",(\S)", r", \1", s)
        for fmt in (
            "%Y-%m-%d",
            "%B %d, %Y",
            "%b %d, %Y",
            "%m/%d/%Y",
            "%m-%d-%Y",
        ):
            try:
                return datetime.strptime(s, fmt).replace(hour=12)
            except ValueError:
                continue
    return None


def _f(v) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").replace("$", "").replace("(", "-").replace(")", ""))
    except ValueError:
        return None


def _right(type_val) -> str:
    if not type_val:
        return "PUT"  # default 0DTE fly book bias
    t = str(type_val).strip().upper()
    if t.startswith("C"):
        return "CALL"
    return "PUT"


def _asset_class(asset: str) -> str:
    a = (asset or "SPX").upper()
    if a in ("ES", "MES", "NQ", "MNQ"):
        return "future_option"
    if a in ("TSLA", "AAPL", "SPY", "QQQ"):
        return "equity_option"
    return "equity_option"


def _strike_tick(asset: str) -> int:
    """Listed strike increments (0DTE / standard chain conventions)."""
    a = (asset or "SPX").upper()
    # Mini / cash-settled small-notional: 1-point strikes common
    if a in ("XSP", "SPY", "QQQ", "IWM"):
        return 1
    # Equity singles
    if a in ("TSLA", "AAPL", "NVDA", "AMZN", "META", "GOOGL"):
        return 5
    # SPX / NDX / ES / NQ butterflies: 5-point strikes
    return 5


def _round_strike(k: float, tick: int) -> int:
    return int(round(float(k) / tick) * tick)


# Yahoo symbols for daily close lookup
_YF_SYMBOL = {
    "SPX": "^GSPC",
    "XSP": "^GSPC",  # scale /10 later
    "ES": "ES=F",
    "MES": "ES=F",
    "NDX": "^NDX",
    "NQ": "NQ=F",
    "MNQ": "NQ=F",
    "TSLA": "TSLA",
    "SPY": "SPY",
    "QQQ": "QQQ",
}


class SpotBook:
    """Daily closes by underlier for realistic synthetic body strikes."""

    def __init__(self) -> None:
        self._by_sym: dict[str, dict[str, float]] = {}  # yf_sym → date ISO → close

    def load(self, assets: set[str], start: datetime, end: datetime) -> None:
        try:
            import yfinance as yf
        except ImportError as e:
            raise SystemExit(
                "yfinance required for spot-relative strikes: pip install yfinance"
            ) from e

        yf_syms = sorted(
            {_YF_SYMBOL.get(a.upper(), a.upper()) for a in assets if a}
        )
        if not yf_syms:
            return
        # pad range for weekends / holidays
        start_s = (start - __import__("datetime").timedelta(days=7)).strftime("%Y-%m-%d")
        end_s = (end + __import__("datetime").timedelta(days=7)).strftime("%Y-%m-%d")
        print(f"Downloading spots {yf_syms} {start_s}→{end_s}…")
        data = yf.download(
            yf_syms,
            start=start_s,
            end=end_s,
            progress=False,
            auto_adjust=True,
            threads=True,
        )
        if data is None or len(data) == 0:
            print("WARNING: empty yfinance download; falling back to static anchors")
            return

        # Normalize Close to {symbol: Series}
        close = data["Close"] if "Close" in data.columns else data
        if hasattr(close, "columns"):
            # multi-ticker
            for sym in yf_syms:
                if sym not in close.columns:
                    # single column weirdness
                    continue
                series = close[sym].dropna()
                self._by_sym[sym] = {
                    idx.strftime("%Y-%m-%d"): float(val)
                    for idx, val in series.items()
                }
        else:
            # single ticker Series
            sym = yf_syms[0]
            series = close.dropna()
            self._by_sym[sym] = {
                idx.strftime("%Y-%m-%d"): float(val) for idx, val in series.items()
            }
        for sym, m in self._by_sym.items():
            print(f"  {sym}: {len(m)} days")

    def spot(self, asset: str, day: datetime) -> float | None:
        a = (asset or "SPX").upper()
        yf_sym = _YF_SYMBOL.get(a, a)
        hist = self._by_sym.get(yf_sym)
        if not hist:
            return None
        # walk back up to 7 calendar days for weekends/holidays
        d = day.date() if isinstance(day, datetime) else day
        for _ in range(8):
            key = d.isoformat()
            if key in hist:
                px = hist[key]
                if a == "XSP":
                    return px / 10.0
                return px
            d = d - __import__("datetime").timedelta(days=1)
        return None


def _fallback_spot(asset: str) -> float:
    a = (asset or "SPX").upper()
    if a in ("ES", "MES"):
        return 5000.0
    if a in ("NQ", "MNQ", "NDX"):
        return 18000.0
    if a == "XSP":
        return 500.0
    if a == "TSLA":
        return 250.0
    return 4500.0  # SPX mid-range 2022–26


def _body_strike(
    asset: str,
    tick: int,
    *,
    right: str,
    day: datetime,
    spots: SpotBook | None,
) -> int:
    """
    Body ≈ 3–5 strikes OTM from that day's spot (Call above, Put below).
    Uses Yahoo daily close when available.
    """
    a = (asset or "SPX").upper()
    spot = spots.spot(a, day) if spots else None
    if spot is None or spot <= 0:
        spot = _fallback_spot(a)

    # 3–5 strikes away (deterministic hash from date for variety, not pure random)
    day_key = day.strftime("%Y%m%d") if isinstance(day, datetime) else str(day)
    n_strikes = 3 + (sum(ord(c) for c in day_key + a) % 3)  # 3, 4, or 5
    offset = n_strikes * tick

    if right == "CALL":
        raw = spot + offset  # OTM call fly above
    else:
        raw = spot - offset  # OTM put fly below

    body = _round_strike(raw, tick)
    # keep positive
    if body < tick * 2:
        body = _round_strike(spot, tick)
    return body


def _legs(
    *,
    right: str,
    width: float | None,
    qty: int,
    under: str,
    expiry: datetime,
    pos: str,
    entry_or_exit: float,
    trade_day: datetime | None = None,
    spots: SpotBook | None = None,
) -> list[dict]:
    """
    Synthetic 1-2-1 butterfly from sheet Width/QTY/Type.

    Width = wing distance in points (body to each wing) → K−W, K, K+W on tick.
    Body K ≈ 3–5 strikes from that day's underlying (Call above, Put below).
    """
    under_u = (under or "SPX").upper()
    tick = _strike_tick(under_u)
    raw_w = float(width) if width is not None and float(width) > 0 else 20.0
    w = max(tick, _round_strike(raw_w, tick))
    day = trade_day or expiry
    mid = _body_strike(under_u, tick, right=right, day=day, spots=spots)
    strikes = [mid - w, mid, mid + w]
    for s in strikes:
        if s % tick != 0:
            raise RuntimeError(f"invalid strike {s} for tick {tick} on {under_u}")

    q = max(1, int(qty or 1))
    exp = expiry.date()
    ac = _asset_class(under_u)
    if pos == "TO_OPEN":
        sides = [("BUY", q), ("SELL", 2 * q), ("BUY", q)]
    else:
        sides = [("SELL", q), ("BUY", 2 * q), ("SELL", q)]
    price = abs(float(entry_or_exit or 0))
    mult = 50 if under_u in ("ES", "MES") else 20 if under_u in ("NQ", "MNQ") else 100
    out = []
    for i, ((side, qn), strike) in enumerate(zip(sides, strikes)):
        out.append(
            {
                "leg_index": i,
                "side": side,
                "quantity": qn,
                "pos_effect": pos,
                "asset_class": ac,
                "underlier": under_u,
                "symbol": under_u,
                "expiry": exp,
                "strike": Decimal(str(int(strike))),
                "option_right": right,
                "multiplier": mult,
                "fill_price": Decimal(str(round(price / 3 if price else 0, 4))),
                "fees": None,
            }
        )
    return out


def ensure_identity(cur, email: str) -> int:
    cur.execute(
        "SELECT identity_id FROM identities WHERE email=%s",
        (email,),
    )
    row = cur.fetchone()
    if not row:
        raise SystemExit(f"No identity for {email}")
    return int(row["identity_id"])


def wipe_xlsx_imports(cur, iid: int) -> None:
    cur.execute(
        """DELETE FROM member_trade_log_trades
           WHERE identity_id=%s AND external_adapter=%s""",
        (iid, ADAPTER),
    )
    # Also wipe all if --replace-all: handled by caller


def ensure_account(cur, iid: int, label: str, broker: str) -> int:
    cur.execute(
        """SELECT id FROM member_trade_log_accounts
           WHERE identity_id=%s AND label=%s""",
        (iid, label),
    )
    row = cur.fetchone()
    if row:
        return int(row["id"])
    cur.execute(
        """INSERT INTO member_trade_log_accounts
           (identity_id, label, broker, currency, status, sort_order)
           VALUES (%s,%s,%s,'USD','active',0)""",
        (iid, label, broker),
    )
    return int(cur.lastrowid)


def insert_trade(
    cur,
    *,
    iid: int,
    account_id: int,
    exec_at: datetime,
    strategy: str,
    net_price: float | None,
    net_side: str | None,
    pnl: float | None,
    legs: list[dict],
    external_order_id: str,
    asset: str,
) -> int:
    cur.execute(
        """INSERT INTO member_trade_log_trades
           (identity_id, account_id, exec_at, asset_class, strategy, order_type,
            net_price, net_side, setup_md, plan_md, rules_md, adherence,
            deviation_md, lesson_md, pnl_amount, external_adapter, external_order_id)
           VALUES (%s,%s,%s,%s,%s,'LMT',%s,%s,'','','','unknown','','',%s,%s,%s)""",
        (
            iid,
            account_id,
            exec_at,
            _asset_class(asset),
            strategy,
            Decimal(str(net_price)) if net_price is not None else None,
            net_side,
            Decimal(str(round(pnl, 4))) if pnl is not None else None,
            ADAPTER,
            external_order_id,
        ),
    )
    tid = int(cur.lastrowid)
    for L in legs:
        cur.execute(
            """INSERT INTO member_trade_log_legs
               (trade_id, identity_id, account_id, leg_index, side, quantity,
                pos_effect, asset_class, underlier, symbol, expiry, strike,
                option_right, multiplier, fill_price, fees)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (
                tid,
                iid,
                account_id,
                L["leg_index"],
                L["side"],
                L["quantity"],
                L["pos_effect"],
                L["asset_class"],
                L["underlier"],
                L["symbol"],
                L["expiry"],
                L["strike"],
                L["option_right"],
                L["multiplier"],
                L["fill_price"],
                L["fees"],
            ),
        )
    return tid


def import_log(
    cur,
    iid: int,
    rows: list,
    *,
    replace_all: bool,
    spots: SpotBook | None = None,
) -> tuple[int, int]:
    if replace_all:
        cur.execute(
            "DELETE FROM member_trade_log_trades WHERE identity_id=%s",
            (iid,),
        )
        # keep accounts; ensure Primary
    else:
        wipe_xlsx_imports(cur, iid)

    account_id = ensure_account(cur, iid, "0DTE Book", "thinkorswim")
    n_open = n_close = 0

    for i, r in enumerate(rows):
        open_d = _dt(r[0])
        close_d = _dt(r[1]) or open_d
        if not open_d:
            continue
        typ = r[2]
        entry = _f(r[3])
        exit_ = _f(r[4])
        width = _f(r[8])
        asset = str(r[9] or "SPX").strip().upper()
        qty = int(_f(r[10]) or 1)
        net_pl = _f(r[11])
        if net_pl is None:
            continue

        right = _right(typ)
        # Guard bad spreadsheet years (e.g. open 2026-05-12 / close 2027-05-12)
        # that used to create year-long OPEN interest in the day book.
        hold_days = (close_d.date() - open_d.date()).days
        if hold_days < 0 or hold_days > 30:
            print(
                f"skip row {i}: open={open_d.date()} close={close_d.date()} "
                f"hold_days={hold_days} (max 30) — not imported"
            )
            continue
        # Same-day 0DTE: open morning / close afternoon for ordering
        open_at = open_d.replace(hour=10, minute=0, second=0)
        close_at = close_d.replace(hour=15, minute=30, second=0)
        if close_at <= open_at:
            close_at = open_at.replace(hour=15, minute=30)

        # Same body for open+close pair (geometry must match)
        leg_kw = dict(
            right=right,
            width=width,
            qty=qty,
            under=asset,
            expiry=close_d,
            trade_day=open_d,
            spots=spots,
        )

        oid = f"xlsx-log-{i:04d}"
        insert_trade(
            cur,
            iid=iid,
            account_id=account_id,
            exec_at=open_at,
            strategy="BUTTERFLY",
            net_price=abs(entry) if entry is not None else None,
            net_side="DEBIT",
            pnl=None,
            legs=_legs(
                **leg_kw,
                pos="TO_OPEN",
                entry_or_exit=entry or 0,
            ),
            external_order_id=f"{oid}-open",
            asset=asset,
        )
        n_open += 1
        insert_trade(
            cur,
            iid=iid,
            account_id=account_id,
            exec_at=close_at,
            strategy="BUTTERFLY",
            net_price=abs(exit_) if exit_ is not None else 0,
            net_side="CREDIT" if (exit_ or 0) >= 0 else "DEBIT",
            pnl=net_pl,
            legs=_legs(
                **leg_kw,
                pos="TO_CLOSE",
                entry_or_exit=exit_ or 0,
            ),
            external_order_id=f"{oid}-close",
            asset=asset,
        )
        n_close += 1

    return n_open, n_close


def load_log_rows(path: Path) -> list:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if "Log" not in wb.sheetnames:
        raise SystemExit(f"No 'Log' sheet in {path}; sheets={wb.sheetnames}")
    ws = wb["Log"]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or r[0] is None:
            continue
        if r[11] is None:
            continue
        rows.append(r)
    wb.close()
    return rows


def clone_identity(cur, from_id: int, to_id: int) -> None:
    """Full replace of target trade log with source."""
    cur.execute(
        "DELETE FROM member_trade_log_trades WHERE identity_id=%s", (to_id,)
    )
    cur.execute(
        "DELETE FROM member_trade_log_accounts WHERE identity_id=%s", (to_id,)
    )
    cur.execute(
        "SELECT * FROM member_trade_log_accounts WHERE identity_id=%s",
        (from_id,),
    )
    acct_map: dict[int, int] = {}
    for a in cur.fetchall():
        cur.execute(
            """INSERT INTO member_trade_log_accounts
               (identity_id, label, broker, broker_label, currency, status,
                badge_color, sort_order, notes_md)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (
                to_id,
                a["label"],
                a["broker"],
                a.get("broker_label"),
                a.get("currency") or "USD",
                a["status"],
                a.get("badge_color"),
                a.get("sort_order") or 0,
                a.get("notes_md"),
            ),
        )
        acct_map[int(a["id"])] = int(cur.lastrowid)

    cur.execute(
        "SELECT * FROM member_trade_log_trades WHERE identity_id=%s ORDER BY id",
        (from_id,),
    )
    for t in cur.fetchall():
        new_acct = acct_map.get(int(t["account_id"]))
        if new_acct is None:
            continue
        cur.execute(
            """INSERT INTO member_trade_log_trades
               (identity_id, account_id, exec_at, asset_class, strategy, order_type,
                net_price, net_side, setup_md, plan_md, rules_md, adherence,
                deviation_md, lesson_md, pnl_amount, external_adapter, external_order_id)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (
                to_id,
                new_acct,
                t["exec_at"],
                t.get("asset_class"),
                t["strategy"],
                t.get("order_type") or "LMT",
                t.get("net_price"),
                t.get("net_side"),
                t.get("setup_md") or "",
                t.get("plan_md") or "",
                t.get("rules_md") or "",
                t.get("adherence") or "unknown",
                t.get("deviation_md") or "",
                t.get("lesson_md") or "",
                t.get("pnl_amount"),
                t.get("external_adapter"),
                t.get("external_order_id"),
            ),
        )
        new_tid = int(cur.lastrowid)
        cur.execute(
            "SELECT * FROM member_trade_log_legs WHERE trade_id=%s ORDER BY leg_index",
            (t["id"],),
        )
        for L in cur.fetchall():
            cur.execute(
                """INSERT INTO member_trade_log_legs
                   (trade_id, identity_id, account_id, leg_index, side, quantity,
                    pos_effect, asset_class, underlier, symbol, expiry, strike,
                    option_right, multiplier, fill_price, fees)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                (
                    new_tid,
                    to_id,
                    new_acct,
                    L.get("leg_index") or 0,
                    L["side"],
                    L["quantity"],
                    L.get("pos_effect"),
                    L.get("asset_class"),
                    L.get("underlier"),
                    L.get("symbol"),
                    L.get("expiry"),
                    L.get("strike"),
                    L.get("option_right"),
                    L.get("multiplier"),
                    L.get("fill_price") or 0,
                    L.get("fees"),
                ),
            )


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--email", default="ernie@fattail.ai")
    ap.add_argument(
        "--also-clone-to",
        default="dev-admin@labs.local,ernie@dudefromearth.com,erniev@mac.com",
        help="Comma-separated emails to clone into after import",
    )
    ap.add_argument(
        "--replace-all",
        action="store_true",
        default=True,
        help="Replace entire trade log for target (default True)",
    )
    args = ap.parse_args()
    if not args.xlsx.exists():
        raise SystemExit(f"File not found: {args.xlsx}")

    rows = load_log_rows(args.xlsx)
    print(f"Loaded {len(rows)} Log rows from {args.xlsx}")

    # Historical spots → body strikes ~3–5 ticks OTM from that day's price
    assets: set[str] = set()
    days: list[datetime] = []
    for r in rows:
        od = _dt(r[0])
        if od:
            days.append(od)
        a = str(r[9] or "SPX").strip().upper()
        if a:
            assets.add(a)
    spots = SpotBook()
    if days:
        spots.load(assets, min(days), max(days))

    with db.transaction() as conn:
        with conn.cursor() as cur:
            iid = ensure_identity(cur, args.email)
            n_open, n_close = import_log(
                cur, iid, rows, replace_all=args.replace_all, spots=spots
            )
            print(f"Imported for {args.email}: {n_open} opens + {n_close} closes")

            cur.execute(
                """SELECT COUNT(*) n, SUM(pnl_amount) s
                   FROM member_trade_log_trades
                   WHERE identity_id=%s AND pnl_amount IS NOT NULL""",
                (iid,),
            )
            r = cur.fetchone()
            print(f"  closes with PnL: {r['n']}, sum={r['s']}")

            for email in [e.strip() for e in args.also_clone_to.split(",") if e.strip()]:
                if email == args.email:
                    continue
                try:
                    tid = ensure_identity(cur, email)
                except SystemExit as e:
                    print(f"  skip clone {email}: {e}")
                    continue
                clone_identity(cur, iid, tid)
                print(f"  cloned → {email}")


if __name__ == "__main__":
    main()
